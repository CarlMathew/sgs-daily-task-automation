from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import pandas as pd




class ExcelModule:
    
    def __init__(self):
        self.worklist_row_height: int = 21
        self.rush_row_height: int = 24
        self.late_row_height: int = 24
        
    
    def style_filter_worklist(self, ws, style_not_late: list[dict]) -> None:
        """
            Fill bg in excel into gray
        """ 
        color_gray: str = "BFBFBF"
        fill_bg_gray = PatternFill(fill_type="solid", start_color=color_gray, end_color=color_gray)

        for index, styles in enumerate(style_not_late):
            row_num: int = index + 2
            sample_num_filter: str = styles["SampleNum_Not_Late"]
            acctnum_filter: str = styles["AcctNum_Not_Late"]
            tat_filter : str= styles["TAT_Not_Late"]
            products_filter: str = styles["Products_Not_Late"]
            collected_filter: str = styles["Collected_Not_Late"]

            if sample_num_filter == "1.0":
                ws[f"A{row_num}"].fill = fill_bg_gray
                ws[f"B{row_num}"].fill = fill_bg_gray
                ws[f"K{row_num}"].fill = fill_bg_gray

            elif acctnum_filter == "1.0":
                ws[f"K{row_num}"].fill = fill_bg_gray
                ws[f"E{row_num}"].fill = fill_bg_gray
            
            elif tat_filter == "1.0":
                ws[f"K{row_num}"].fill = fill_bg_gray
                ws[f"H{row_num}"].fill = fill_bg_gray
            
            elif products_filter == "1.0":
                ws[f"K{row_num}"].fill = fill_bg_gray
                ws[f"L{row_num}"].fill = fill_bg_gray
            
            elif collected_filter == "1.0":
                ws[f"K{row_num}"].fill = fill_bg_gray
                ws[f"J{row_num}"].fill = fill_bg_gray



    def process_automation_worklist_dayton(self, ws, df) -> bool:
        """
            For Dayton automation
            Start the automation of copy values from excel to excel on worklist data
        """

        ws.auto_filter.ref = f"A1:N1"
        
        additional_columns: list[str] = ["SampleNum_Not_Late", "AcctNum_Not_Late", "TAT_Not_Late", "Products_Not_Late", "Collected_Not_Late"]
        data_of_worklist = df.drop(columns = additional_columns)
        filter_of_worklist = df[additional_columns].to_dict(orient="records")

        df_to_dict: list[dict] = data_of_worklist.to_dict(orient="records")
        columns: list[str] = tuple(df_to_dict[0].keys())
        data: list[tuple] = [tuple(data.values()) for data in df_to_dict]
        data.insert(0, columns)


        self.style_filter_worklist(ws, filter_of_worklist)

        try:

            for row in range(1, len(data) + 1):
                ws.row_dimensions[row].height = self.worklist_row_height

                for col in range(len(columns)):
                    char: str = get_column_letter(col + 1) 

                    cell: str = char + str(row)
                    value: str = data[row-1][col]

                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

                    if char == "C":
                        try:
                            split_1_value: str | int = int(value)
                            ws[cell] = split_1_value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)
                    elif char == "H":
                        ws[cell] = value
                        ws[cell].number_format = "0"
                    elif char == "K":
                        # date_value: datetime = datetime.strptime(value, "%Y-%m-%d")
                        ws[cell] = value
                        ws[cell].number_format = numbers.FORMAT_DATE_DDMMYY
                    elif char == "N":
                        try:
                            comments_eta_value: str | int = int(value)
                            ws[cell] = comments_eta_value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)

                    else:
                        try:
                            value: str | int = int(value)
                            ws[cell] = value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)
            return True
        except Exception as e:
            return False
                

                


    def process_automation_worklist(self, ws, columns:list[str], data:list[tuple]) -> bool:
        """
            Start the automation of copy values from excel to excel on worklist data
        """
        try:
            for row in range(1, len(data) + 1):
                ws.row_dimensions[row].height = self.worklist_row_height
                for col in range(len(columns)):
                    char: str = get_column_letter(col + 1)
                    
                    cell: str = char + str(row)
                    value:str = data[row -1][col]
                    ws[cell].alignment = Alignment(horizontal="center", vertical="center")
                    

                    if char == "C":
                        try:
                            split_1_value: str | int = int(value)
                            ws[cell] = split_1_value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)
                    elif char == "H":
                        ws[cell] = value
                        ws[cell].number_format = "0"
                    elif char == "K":
                        # date_value: datetime = datetime.strptime(value, "%Y-%m-%d")
                        ws[cell] = value
                        ws[cell].number_format = numbers.FORMAT_DATE_DDMMYY
                    elif char == "N":
                        try:
                            comments_eta_value: str | int = int(value)
                            ws[cell] = comments_eta_value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)

                    else:
                        try:
                            value: str | int = int(value)
                            ws[cell] = value
                            ws[cell].number_format = numbers.FORMAT_TEXT
                        except:
                            ws[cell] = str(value)

            return True
        except Exception as e:
            return False

    def process_automation_rush(self, ws, df) -> int:
        """Start the automation of copy values from excel to excel on rush data"""
        font_and_cell_color_df: pd.DataFrame = df[["Font", "FillColor"]]
        remaining_columns_df: pd.DataFrame = df.drop(columns = ["Font", "FillColor"])

       
        remaining_columns_df["Comments/ETA"] = remaining_columns_df["Comments/ETA"].astype(str)

        font_and_cell_color_dict: list[dict] = font_and_cell_color_df.to_dict(orient="records")
        remaining_columns_dict: list[dict] = remaining_columns_df.to_dict(orient = "records")
        
        self.change_style_rush(ws, df)
        
        columns: list[str] = list(remaining_columns_df.columns)
        rows_data: list[tuple] = [tuple(rows.values()) for rows in remaining_columns_dict]

        rows_data.insert(0, columns)
        
        # for data, styles in zip(rows_data, font_and_cell_color_dict):
        for row in range(1, len(rows_data) + 1):
            for col in range(len(columns)):
                char: str = get_column_letter(col + 1)
                cell_name: str = char + str(row)
                cell_value:str = rows_data[row - 1][col]

                
                ws[cell_name].alignment = Alignment(horizontal="center", vertical="center")

                if char == "G":
                    try: 
                        number_value: int = round(float(cell_value))
                        ws[cell_name] = number_value
                    except:
                        string_value: str = cell_value
                        ws[cell_name] = string_value
                elif char == "H" or char == "I":
                    try:
                        number_value: int = int(cell_value)
                        ws[cell_name] = number_value
                    except:
                        string_value: str = str(cell_value)
                        ws[cell_name] = string_value
                elif char == "L": 
                    ws[cell_name].alignment = Alignment(horizontal="center", vertical="center")
                    string_value: str = cell_value
                    ws[cell_name] = string_value
                elif char == "M":
                    ws[cell_name].alignment = Alignment(horizontal="center", vertical="center")
                    try:
                        number_value: int = round(float(cell_value))
                        ws[cell_name] = number_value
                    except:
                        if cell_value == "nan":
                            no_value: str = " "
                            ws[cell_name] = no_value
                        else:
                            ws[cell_name] = cell_value
                else:
                    ws[cell_name] = cell_value
        
        # Adding table on rush
        max_column:int = ws.max_column
        last_col_letter:str = ws.cell(row=1, column=max_column).column_letter
        table_rush_range:str = f"A1:{last_col_letter}{len(rows_data)}"

        table_ref = Table(displayName="Rush_Table", ref=table_rush_range)

        # style_table = TableStyleInfo(name="TableStyleLight1")

        # table_ref.tableStyleInfo = style_table
        
        ws.add_table(table_ref)


        # Adding Features    
        return len(rows_data)

    def adding_last_column_on_rush(self, ws, last_three_columns: pd.DataFrame, last_row_num_of_data: int, column_num: int) -> None:
        """Add the last row of the data into excel"""

        data: list[dict]  = last_three_columns[last_three_columns.columns[:column_num]].to_dict(orient="records")[1:]

        data: list[tuple]= [tuple(datum.values()) for datum in data]


        for row in range(len(data)):

            starting_row =(last_row_num_of_data + 2)  + row
            for col in range(1, column_num + 1):
                char = get_column_letter(col)
                value = data[row][col - 1]


                try:
                    ws[f"{char}{starting_row}"] = int(value)
                except Exception as e:
                    ws[f"{char}{starting_row}"] = value



    def change_style_rush(self, ws, df) -> None:
        """ Update the font and fill color of each cell (Rush)"""
        dataframe_dict: list[dict] = df.to_dict(orient = "records")
        length_of_columns: int = len(df.columns) - 2
        length_of_rows: int = len(dataframe_dict)


        for row in range(2, length_of_rows + 2):
            ws.row_dimensions[row].height = self.rush_row_height
            fonts: str = dataframe_dict[row - 2]["Font"]
            fill_color:str = dataframe_dict[row-2]["FillColor"]
            

            for col in range(length_of_columns):
                char: str = get_column_letter(col + 1)
                if fonts == "RedAndBold":
                    red_bold_font = Font(color='FF0000', bold=True)
                    ws[char + str(row)].font = red_bold_font
                    

                fill_bg = PatternFill(fill_type="solid", fgColor=fill_color)
                ws[char + str(row)].fill = fill_bg



    def process_automation_late_repgen(self,ws,df:pd.DataFrame, report_type:str) -> None:

        """Start the automation of copy values from excel to excel on late and repgen data """
        # font_and_cell_color_df: pd.DataFrame = df[["Fill Color", "Font"]]
        remaining_columns_df: pd.DataFrame = df.drop(columns=["Font", "Fill Color"])

        remaining_columns_dict: list[dict] = remaining_columns_df.to_dict(orient = "records")

        self.change_style_late(ws, df)

        columns: list[str] = list(remaining_columns_df.columns)
        rows_data: list[tuple] = [tuple(rows.values()) for rows in remaining_columns_dict]

        rows_data.insert(0, columns)


        for row in range(1, len(rows_data) + 1):
            for col in range(len(columns)):
                char: str = get_column_letter(col + 1)
                cell_name: str = char + str(row)
                cell_value:str = rows_data[row - 1][col]

                
                ws[cell_name].alignment = Alignment(horizontal="center", vertical="center")


                if char == "A":
                    ws[cell_name].alignment = Alignment(horizontal="left", vertical="center")

                elif char == "H" or char == "I":
                    try:
                        number_value: int = int(cell_value)
                        ws[cell_name] = number_value
                    except:
                        string_value: str = str(cell_value)
                        ws[cell_name] = string_value
                
                elif char == "N":
                    try:
                        number_value: int = int(cell_value)
                        ws[cell_name] = number_value
                    except:
                        string_value: str = str(cell_value)
                        ws[cell_name] = string_value


                ws[cell_name] = cell_value


        if report_type == "LATE":
            max_column:int = ws.max_column
            last_col_letter:str = ws.cell(row=1, column=max_column).column_letter
            table_rush_range:str = f"A1:{last_col_letter}{len(rows_data)}"

            table_ref = Table(displayName="Late_Table", ref=table_rush_range)

            style_table = TableStyleInfo(name="TableStyleMedium3")

            table_ref.tableStyleInfo = style_table
            
            ws.add_table(table_ref)

        elif report_type == "REPGEN":
            max_column = ws.max_column
            last_col_letter = ws.cell(row=1, column=max_column).column_letter
            ws.auto_filter.ref = f"A1:{last_col_letter}1"

        return len(rows_data)
    

    def change_style_late(self, ws, df) -> None:
        """Update the font and fill color of each cell (Late)"""
        dataframe_dict: list[dict] = df.to_dict(orient = "records")
        length_of_columns: int = len(df.columns) - 2
        length_of_rows: int = len(dataframe_dict)


        for row in range(2, length_of_rows + 2):
            ws.row_dimensions[row].height = self.late_row_height
            fonts: str = dataframe_dict[row - 2]["Font"]
            fill_color:str = dataframe_dict[row-2]["Fill Color"]

            for col in range(length_of_columns):
                char:str = get_column_letter(col + 1)

                if fonts == "RedAndBold":
                    red_bold_font = Font(color='FF0000', bold=True)
                    ws[char + str(row)].font = red_bold_font
                elif fonts == "PurpleAndBold":
                    purple_bold_font = Font(color='7030A0', bold=True)
                    ws[char + str(row)].font = purple_bold_font

                if pd.isna(fill_color):
                    pass
                else:
                    fill_bg = PatternFill(fill_type="solid", fgColor=fill_color)
                    ws[char + str(row)].fill = fill_bg




    def insert_data_from_template(self, 
                                  files: list[str],
                                  from_excel:str="",
                                  to_excel:str="",
                                  automation_type:int=1,
                                  site:dict={},
                                  last_three_columns:pd.DataFrame = "") -> None :
        """
            Transfer data from excel file to excel file
        """
        wb = load_workbook(from_excel)

        for file in files:
            if automation_type == 1:
                # Worklist automation V1 
                report_type: str = site.get(os.path.split(file)[-1].split("_")[0], "")
                df = pd.read_csv(
                    file,
                    sep="\t",
                    engine="python"

                )
                ws = wb[report_type]
                
                max_column = ws.max_column
                last_col_letter = ws.cell(row=1, column=max_column).column_letter
                ws.auto_filter.ref = f"A1:{last_col_letter}1"

                df_to_dict: list[dict] = df.to_dict(orient="records")
                columns: list[str] = tuple(df_to_dict[0].keys())
                data: list[tuple] = [tuple(data.values()) for data in df_to_dict]
                data.insert(0, columns)
                isFinished:bool = self.process_automation_worklist(ws, columns=columns, data=data)

                wb.save(to_excel)


            elif automation_type == 2:
                # Rush Prio Automation
                report_type = "RUSH_PRIORITY"
                df = pd.read_csv(
                    file,
                    sep="\t",
                    engine="python"

                )
                ws = wb[report_type]

                last_row_num_of_data: int = self.process_automation_rush(ws, df)

                # Still on Testing. Do not remove the try if not yet stable
                try:
                    self.adding_last_column_on_rush(ws, last_three_columns, last_row_num_of_data, 2)
                except Exception as e:
                    print(f'Error: {e}')
            
                wb.save(to_excel)
            
            elif automation_type == 3:
                # Late Automation

                df = pd.read_csv(file, sep="\t", engine="python")

                ws = wb[site]

                last_row_num_of_data = self.process_automation_late_repgen(ws, df, site)
                # Still on Testing. Do not remove the try if not yet stable
                try:
                    self.adding_last_column_on_rush(ws, last_three_columns, last_row_num_of_data, 4)
                except Exception as e:
                    print(f'Error: {e}')
            

                wb.save(to_excel)
            
            elif automation_type == 4:

                # Repgen Automation
                df = pd.read_csv(file)

                ws = wb[site]

                last_row_num_of_data = self.process_automation_late_repgen(ws, df, site)
                # Still on Testing. Do not remove the try if not yet stable
                try:
                    self.adding_last_column_on_rush(ws, last_three_columns, last_row_num_of_data, 4)
                except Exception as e:
                    print(f'Error: {e}')

                wb.save(to_excel)
            
            elif automation_type == 5:
                # Worklist Automation Dayton V2
                df = pd.read_csv(file, sep = "\t", engine = "python")
                report_type: str = site.get(os.path.split(file)[-1].split("_")[0], "")
                ws = wb[report_type]
                self.process_automation_worklist_dayton(ws, df)
                wb.save(to_excel)

        wb.close()




